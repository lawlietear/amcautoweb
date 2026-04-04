# -*- coding: utf-8 -*-
"""
Excel 处理模块
支持读取、验证、状态更新
"""

import pandas as pd
import os


class ExcelHandler:
    """Excel 处理类"""
    
    # 标准列定义（支持新旧字段名）
    REQUIRED_COLS = [
        # 核心字段（必须存在）
        '批复编号', '债务人名称',
        '债权本金金额', '债权总额',
        # 其他字段（新旧兼容）
        '生成状态', '户数', '项目名称', '部门',
        '尽调报告名称', '报告名称',  # 新旧名称兼容
        '项目负责人', '项目负责人名称',  # 新旧名称兼容
        '会议/审批日', '会议时间',  # 新旧名称兼容
        '会次', '会议次数',  # 新旧名称兼容
        '业务分类', '基准日',
        '利息金额', '债权利息金额',  # 新旧名称兼容
        '其他费用金额',
        '拟转让价格（元）', '拟转让价格',  # 新旧名称兼容
        '交易费用承担主体',
        '保证金金额（元）', '保证金金额',  # 新旧名称兼容
        '剩余交易价款金额（元）',
        '尾款支付期限',
        '资金占用费率',
        '审核人员'
    ]
    
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.df = None
    
    def read(self):
        """
        读取 Excel 文件
        注意：大文件(>10MB)建议使用 chunksize 参数分批读取
        当前场景（单次上传<16MB）同步读取可接受
        """
        import logging
        logger = logging.getLogger('autodocweb')

        try:
            if not os.path.exists(self.excel_path):
                logger.error(f"Excel 文件不存在：{self.excel_path}")
                return False

            # 对于大文件可考虑: pd.read_excel(self.excel_path, chunksize=1000)
            self.df = pd.read_excel(self.excel_path)

            # 如果没有状态列，自动添加
            if '生成状态' not in self.df.columns:
                self.df['生成状态'] = ""
                logger.warning(f"已自动添加'生成状态'列")

            return True
        except Exception as e:
            logger.error(f"读取 Excel 失败：{e}")
            return False
    
    def validate_structure(self):
        """
        验证 Excel 列名是否完整
        只检查核心字段，其他字段可选
        """
        if self.df is None:
            return False
        
        # 核心必须字段
        CORE_COLS = ['批复编号', '债务人名称', '债权本金金额', '债权总额']
        
        missing = [col for col in CORE_COLS if col not in self.df.columns]
        if missing:
            print(f"❌ Excel 缺少核心列：{', '.join(missing)}")
            print(f"现有列：{list(self.df.columns)}")
            return False
        
        # 检查是否至少有部分字段匹配
        matched_cols = [col for col in self.REQUIRED_COLS if col in self.df.columns]
        if len(matched_cols) < 5:
            print(f"⚠️ Excel 列匹配度较低，请检查列名是否符合要求")
            print(f"现有列：{list(self.df.columns)}")
            return False
        
        print(f"✅ Excel 验证通过，共匹配 {len(matched_cols)} 个字段")
        return True
    
    def get_pending_records(self):
        """
        获取待生成的记录（状态为'立即生成'）
        """
        if self.df is None:
            return None
        
        # 筛选待生成记录（排除空行和表头行）
        pending = self.df[
            (self.df['生成状态'].astype(str) == "立即生成") & 
            (self.df['批复编号'].notna()) &
            (self.df['批复编号'].astype(str) != "批复编号")  # 排除表头行
        ]
        
        return pending
    
    def update_status(self, project_no, new_status="已生成"):
        """
        更新指定项目的生成状态 - 使用 openpyxl 增量更新避免全量读写
        失败时回退到全量读写模式
        """
        import logging
        logger = logging.getLogger('autodocweb')

        # 先尝试增量更新
        if self._update_status_incremental(project_no, new_status):
            return True

        # 失败时回退到全量模式
        logger.warning("增量更新失败，回退到全量读写模式")
        return self._update_status_full(project_no, new_status)

    def _update_status_incremental(self, project_no, new_status):
        """使用 openpyxl 进行增量更新"""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.excel_path)
            ws = wb.active

            # 找到列索引
            project_col = None
            status_col = None

            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value == '批复编号':
                    project_col = col_idx
                elif cell.value == '生成状态':
                    status_col = col_idx

            if not project_col:
                return False

            # 查找并更新
            for row_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row_idx, project_col).value
                if str(cell_value) == str(project_no):
                    if status_col:
                        ws.cell(row_idx, status_col, new_status)
                    else:
                        # 没有状态列则添加到最后一列
                        ws.cell(row_idx, ws.max_column + 1, new_status)
                        ws.cell(1, ws.max_column, '生成状态')
                    wb.save(self.excel_path)
                    return True

            return False

        except Exception:
            return False

    def _update_status_full(self, project_no, new_status):
        """全量读写模式（作为 fallback）"""
        import logging
        logger = logging.getLogger('autodocweb')

        try:
            df = pd.read_excel(self.excel_path)

            mask = df['批复编号'].astype(str) == str(project_no)
            if mask.any():
                df.loc[mask, '生成状态'] = new_status
                df.to_excel(self.excel_path, index=False, engine='openpyxl')
                return True
        except Exception as e:
            logger.warning(f"全量更新 Excel 状态失败：{e}")

        return False
    
    def find_by_project_no(self, project_no):
        """
        根据项目编号查找记录（支持模糊匹配）
        """
        if self.df is None:
            return None
        
        # 模糊匹配
        mask = self.df['批复编号'].astype(str).str.contains(project_no, na=False)
        if mask.any():
            return self.df[mask].iloc[0]
        
        return None
    
    def get_row_data(self, idx):
        """
        获取指定索引的行数据
        """
        if self.df is None or idx >= len(self.df):
            return None
        return self.df.iloc[idx]
